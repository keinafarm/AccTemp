# -*- coding: utf-8 -*-
import openpyxl
import datetime

"""

    Excel File操作部

"""


class NoneCell:
    """
    valueに何をセットされても、常にNoneな値を返すCell
    """

    @property
    def value(self):
        return None

    @value.setter
    def value(self, arg):
        pass


class ATCell:
    """
    valueが抱えているcellのvalueに見えるクラス
    """

    def __init__(self, cell):
        self.cell = cell

    @property
    def value(self):
        return self.cell.value

    @value.setter
    def value(self, value):
        self.cell.value = value


class ATRow:
    """
    Excelのセルデータと、本システム内部の管理情報の構造を橋渡しする
    """

    def __init__(self, row_data, column_position):
        self.row = row_data
        self.column_position = column_position
        self.cell_data = {}
        for key in column_position.keys():
            if column_position[key] is None:
                cell = NoneCell()  # 該当カラムが存在しない時は、常にNoneを返す
            else:
                cell = row_data[column_position[key] - 1]
            cell_object = ATCell(cell)
            self.cell_data[key] = cell_object

    @property
    def start_date(self):
        return self.cell_data["開始日"].value

    @start_date.setter
    def start_date(self, value):
        self.cell_data["開始日"].value = value

    @property
    def end_date(self):
        return self.cell_data["終了日"].value

    @end_date.setter
    def end_date(self, value):
        self.cell_data["終了日"].value = value

    @property
    def target_temperature(self):
        return self.cell_data["目標積算温度"].value

    @target_temperature.setter
    def target_temperature(self, value):
        self.cell_data["目標積算温度"].value = value

    @property
    def current_temperature(self):
        return self.cell_data["現状積算温度"].value

    @current_temperature.setter
    def current_temperature(self, value):
        self.cell_data["現状積算温度"].value = value

    @property
    def rate(self):
        return self.cell_data["目標到達度"].value

    @rate.setter
    def rate(self, value):
        self.cell_data["目標到達度"].value = value

    @property
    def estimate_date(self):
        return self.cell_data["予測終了日"].value

    @estimate_date.setter
    def estimate_date(self, value):
        self.cell_data["予測終了日"].value = value

    def rowNo(self):
        return self.row[0].row

    def __str__(self):
        result = "開始日:" + (
            "空白" if self.start_date is None else self.start_date.strftime('%m/%d'))
        result += "  終了日:" + (
            "空白" if self.end_date is None else self.end_date.strftime('%m/%d'))
        result += "  目標積算温度:" + (
            "空白" if self.target_temperature is None else str(self.target_temperature))
        result += "  現状積算温度:" + (
            "空白" if self.current_temperature is None else str(self.current_temperature))
        result += "  目標到達度:" + (
            "空白" if self.rate is None else str(self.rate))
        result += "  予測終了日:" + (
            "空白" if self.estimate_date is None else self.estimate_date.strftime(
                '%m/%d'))
        return result


class ATSheet:
    def __init__(self, work_sheet):
        """
        積算温度シートクラス

        ２行目に、開始日", "終了日", "目標積算温度", "現状積算温度", "目標到達度", "予測終了日"の各タイトルが出てくるカラム位置を記憶する

        :param work_sheet: ワークシート
        """
        self.work_sheet = work_sheet
        self.column_position = {"開始日": None, "終了日": None, "目標積算温度": None, "現状積算温度": None, "目標到達度": None, "予測終了日": None}
        for cell in work_sheet.iter_cols(min_row=2, max_row=2):
            compare_list = ["開始日", "終了日", "目標積算温度", "現状積算温度", "目標到達度", "予測終了日"]
            value = cell[0].value
            if value in compare_list:
                self.column_position[cell[0].value] = cell[0].column

        self.rows = []
        for row_data in work_sheet.iter_rows(min_row=3):
            row_obj = ATRow(row_data, self.column_position)
            self.rows.append(row_obj)

    @property
    def data_list(self):
        return self.rows


class ATWorkBook:
    def __init__(self, file_name):
        """
        積算温度ワークブッククラス
        :param file_name: エクセルファイル名
        """
        self.file_name = file_name
        self.work_book = openpyxl.load_workbook(file_name)
        """
    開始日(start_date)
    終了日(end_date)
    目標積算温度(target_temperature)
    現状積算温度(current_temperature)
    目標到達度(rate)
    予測終了日(estimate_date)
        """
        self.target_sheets = []
        for ws in self.work_book:  # work sheetの名前リストでループ
            if ws.title.find("積算温度") == 0:
                ws_obj = ATSheet(ws)
                self.target_sheets.append(ws_obj)

        print(self.target_sheets)

    def flash(self):
        self.work_book.save(self.file_name)

    @property
    def sheets_list(self):
        return self.target_sheets


if __name__ == "__main__":
    obj = ATWorkBook("ZGIS_data7.xlsx")
    for sheet in obj.sheets_list:
        for data in sheet.data_list:
            print(data)

    row = obj.sheets_list[0].data_list[2]
    row.target_temperature = 256
    row.current_temperature = 512
    row.rate = 20.3
    row.estimate_date = datetime.datetime(2020, 12, 24)

    row = obj.sheets_list[0].data_list[0]
    row.start_date = datetime.datetime(2020, 2, 24)
    row.end_date = datetime.datetime(2020, 3, 24)
    obj.flash()
    print("Done")
